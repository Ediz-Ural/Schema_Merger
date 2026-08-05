"""End-to-end MVP flow: analyze (LLM) -> user approval -> apply (no LLM)."""

from __future__ import annotations

import dataclasses
import json
from pathlib import Path

from openpyxl import load_workbook

from cli.main import REVIEW_GUARD_EXIT_CODE, main
from core.contracts import MappingContract, dump_mapping, load_mapping
from core.llm import FakeLLMClient


SCHEMA_YAML = """target_columns:
  - name: product_name
    type: string
    required: true
  - name: unit_price
    type: decimal
    required: true
  - name: sold_on
    type: date
    required: true
output:
  format: xlsx
  add_provenance: true
"""


def _llm_response(product: str, price: str, sold_on: str, price_confidence: float) -> str:
    return json.dumps(
        {
            "matches": [
                {
                    "target_column": "product_name",
                    "column": product,
                    "confidence": 0.95,
                    "reason": "Names and samples agree.",
                },
                {
                    "target_column": "unit_price",
                    "column": price,
                    "confidence": price_confidence,
                    "reason": "Decimal samples agree.",
                },
                {
                    "target_column": "sold_on",
                    "column": sold_on,
                    "confidence": 0.93,
                    "reason": "Both columns hold dates.",
                },
            ]
        }
    )


def _approve_reviews(mapping_path: Path) -> int:
    """Stand in for the user editing mapping.yaml: accept every review row."""

    mapping = load_mapping(mapping_path)
    approved = 0
    entries = []
    for entry in mapping.entries:
        sources = []
        for source in entry.sources:
            if source.status == "review":
                approved += 1
                source = dataclasses.replace(source, status="auto", confidence=1.0)
            sources.append(source)
        entries.append(dataclasses.replace(entry, sources=sources))
    dump_mapping(MappingContract(entries=entries), mapping_path)
    return approved


def test_analyze_then_approve_then_apply_produces_the_merged_table(tmp_path, capsys):
    sales = tmp_path / "sales_2023.csv"
    sales.write_text("urun;fiyat;tarih\nKalem;12,50;31.01.2024\nDefter;1.234,56;01.02.2024\n", encoding="utf-8")
    export = tmp_path / "export_q4.csv"
    export.write_text("product,price,date\nPencil,8.90,2024-03-15\n", encoding="utf-8")
    schema = tmp_path / "schema.yaml"
    schema.write_text(SCHEMA_YAML, encoding="utf-8")
    mapping_path = tmp_path / "mapping.yaml"
    merged = tmp_path / "merged.xlsx"

    # --- Phase 1: analyze (the only phase that talks to an LLM) ---------------
    client = FakeLLMClient(
        responses=[
            _llm_response("urun", "fiyat", "tarih", 0.97),
            _llm_response("product", "price", "date", 0.55),  # low -> review
        ]
    )
    analyze_code = main(
        [
            "analyze",
            "--inputs",
            str(sales),
            str(export),
            "--target-schema",
            str(schema),
            "--out",
            str(mapping_path),
        ],
        llm_client=client,
    )
    assert analyze_code == 0
    assert len(client.calls or []) == 2  # one metadata request per file, not per row
    assert not merged.exists()  # analyze plans, it never merges

    # --- The review-guard blocks apply before the user approves --------------
    guard_code = main(["apply", "--mapping", str(mapping_path), "--out", str(merged), "--format", "xlsx"])
    assert guard_code == REVIEW_GUARD_EXIT_CODE
    assert not merged.exists()
    assert "unit_price ← export_q4.csv:price" in capsys.readouterr().out

    # --- The user approves the plan ------------------------------------------
    assert _approve_reviews(mapping_path) == 1

    # --- Phase 2: apply (deterministic, no LLM) ------------------------------
    apply_code = main(["apply", "--mapping", str(mapping_path), "--out", str(merged), "--format", "xlsx"])
    assert apply_code == 0
    assert len(client.calls or []) == 2  # unchanged: apply made no LLM call

    sheet = load_workbook(merged)["merged"]
    header = [cell.value for cell in sheet[1]]
    rows = [dict(zip(header, (cell.value for cell in row))) for row in sheet.iter_rows(min_row=2)]
    assert len(rows) == 3
    assert [row["product_name"] for row in rows] == ["Kalem", "Defter", "Pencil"]
    assert [row["unit_price"] for row in rows] == [12.5, 1234.56, 8.9]
    assert [row["sold_on"].date().isoformat() for row in rows] == [
        "2024-01-31",
        "2024-02-01",
        "2024-03-15",
    ]
    assert [row["_source_file"] for row in rows] == ["sales_2023.csv", "sales_2023.csv", "export_q4.csv"]
    assert [row["unit_price_source_column"] for row in rows] == ["fiyat", "fiyat", "price"]

    report = tmp_path / "merge_report.xlsx"
    assert report.is_file()
    output = capsys.readouterr().out
    assert f"3 satır yazıldı: {merged}" in output
    assert str(report) in output
