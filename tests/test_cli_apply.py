from __future__ import annotations

from pathlib import Path
import sqlite3

import polars as pl
import pytest
from openpyxl import load_workbook

from cli.main import REVIEW_GUARD_EXIT_CODE, main


SCHEMA_YAML = """target_columns:
  - name: product_name
    type: string
    required: true
  - name: unit_price
    type: decimal
    required: true
  - name: stock_quantity
    type: integer
    required: false
output:
  format: csv
  add_provenance: true
"""


def _sources_block(column_tr: str | None, column_en: str | None, status: str, confidence: float) -> str:
    """One ``sources`` block covering both input files of the workspace."""

    def source(file_name: str, column: str | None) -> str:
        rendered = "null" if column is None else column
        return (
            f"    - file: {file_name}\n"
            f"      column: {rendered}\n"
            f"      confidence: {confidence}\n"
            f"      status: {status}\n"
            f"      reason: test\n"
        )

    return source("sales_tr.csv", column_tr) + source("export_en.csv", column_en)


def _workspace(tmp_path: Path, *, price_status: str = "auto", stock_matched: bool = True) -> Path:
    """A self-contained folder: two sources, a target schema, and a plan."""

    (tmp_path / "sales_tr.csv").write_text(
        "urun;fiyat;stok\nKalem;12,50;10\nDefter;1.234,56;3\n", encoding="utf-8"
    )
    (tmp_path / "export_en.csv").write_text(
        "product,price,stock\nPencil,8.90,7\n", encoding="utf-8"
    )
    (tmp_path / "schema.yaml").write_text(SCHEMA_YAML, encoding="utf-8")

    if stock_matched:
        stock = _sources_block("stok", "stock", "auto", 1.0)
    else:
        stock = _sources_block(None, None, "unmatched", 0.0)
    (tmp_path / "mapping.yaml").write_text(
        "- target_column: product_name\n  sources:\n"
        + _sources_block("urun", "product", "auto", 0.99)
        + "- target_column: unit_price\n  sources:\n"
        + _sources_block("fiyat", "price", price_status, 0.54 if price_status == "review" else 0.98)
        + "- target_column: stock_quantity\n  sources:\n"
        + stock,
        encoding="utf-8",
    )
    return tmp_path


def _apply_argv(workspace: Path, out: Path, output_format: str | None = None) -> list[str]:
    argv = ["apply", "--mapping", str(workspace / "mapping.yaml"), "--out", str(out)]
    if output_format is not None:
        argv += ["--format", output_format]
    return argv


def test_apply_merges_and_reports_when_every_match_is_auto(tmp_path, capsys):
    workspace = _workspace(tmp_path)
    out = tmp_path / "out" / "merged.csv"

    exit_code = main(_apply_argv(workspace, out, "csv"))

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame.height == 3
    assert frame["product_name"].to_list() == ["Kalem", "Defter", "Pencil"]
    assert frame["unit_price"].to_list() == [12.5, 1234.56, 8.9]
    assert frame["stock_quantity"].to_list() == [10, 3, 7]
    # Provenance is always present (spec §6/§14).
    assert frame["_source_file"].to_list() == ["sales_tr.csv", "sales_tr.csv", "export_en.csv"]
    assert frame["unit_price_source_column"].to_list() == ["fiyat", "fiyat", "price"]

    report = out.parent / "merge_report.xlsx"
    assert report.is_file()
    summary = {row[0].value: row[1].value for row in load_workbook(report)["Summary"].iter_rows(min_row=2)}
    assert summary["merged_rows"] == 3

    output = capsys.readouterr().out
    assert f"3 satır yazıldı: {out}" in output
    assert "0 boş hücre (null)" in output
    assert str(report) in output


def test_apply_stops_and_writes_nothing_while_a_match_is_in_review(tmp_path, capsys):
    workspace = _workspace(tmp_path, price_status="review")
    out = tmp_path / "merged.csv"

    exit_code = main(_apply_argv(workspace, out, "csv"))

    assert exit_code == REVIEW_GUARD_EXIT_CODE
    assert not out.exists()
    assert not (tmp_path / "merge_report.xlsx").exists()

    output = capsys.readouterr().out
    assert "apply durdu" in output
    assert "2 eşleştirme hâlâ onay bekliyor (review)" in output
    # The blocking columns are listed by name so the user knows what to fix.
    assert "unit_price ← sales_tr.csv:fiyat" in output
    assert "unit_price ← export_en.csv:price" in output
    assert "product_name" not in output.split("apply durdu")[1].split("→")[0]


def test_unmatched_column_becomes_null_without_losing_rows(tmp_path, capsys):
    workspace = _workspace(tmp_path, stock_matched=False)
    out = tmp_path / "merged.csv"

    exit_code = main(_apply_argv(workspace, out, "csv"))

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame.height == 3  # no row is dropped
    assert frame["stock_quantity"].null_count() == 3
    assert frame["stock_quantity_source_column"].null_count() == 3
    assert "3 boş hücre (null)" in capsys.readouterr().out


@pytest.mark.parametrize("output_format", ["csv", "xlsx", "sql"])
def test_each_format_produces_the_matching_output(tmp_path, output_format):
    workspace = _workspace(tmp_path)
    out = tmp_path / f"merged.{output_format}"

    assert main(_apply_argv(workspace, out, output_format)) == 0
    assert out.is_file()
    assert (tmp_path / "merge_report.xlsx").is_file()

    if output_format == "csv":
        assert pl.read_csv(out).height == 3
    elif output_format == "xlsx":
        sheet = load_workbook(out)["merged"]
        assert sheet.max_row == 4  # header + 3 rows
        assert [cell.value for cell in sheet[1]][:3] == [
            "product_name",
            "unit_price",
            "stock_quantity",
        ]
    else:
        script = out.read_text(encoding="utf-8")
        connection = sqlite3.connect(":memory:")
        try:
            connection.executescript(script)
            rows = connection.execute(
                'SELECT product_name, unit_price, "_source_file" FROM merged ORDER BY rowid'
            ).fetchall()
        finally:
            connection.close()
        assert [row[0] for row in rows] == ["Kalem", "Defter", "Pencil"]
        assert rows[2][2] == "export_en.csv"


def test_format_defaults_to_the_target_schema_setting(tmp_path):
    workspace = _workspace(tmp_path)
    out = tmp_path / "merged.csv"

    assert main(_apply_argv(workspace, out)) == 0  # schema.yaml says format: csv

    assert pl.read_csv(out).height == 3


def test_apply_reports_a_missing_source_file_without_writing_output(tmp_path, capsys):
    workspace = _workspace(tmp_path)
    (workspace / "export_en.csv").unlink()
    out = tmp_path / "merged.csv"

    exit_code = main(_apply_argv(workspace, out, "csv"))

    assert exit_code == 2
    assert not out.exists()
    assert "Kaynak dosya bulunamadı: export_en.csv" in capsys.readouterr().out


def test_apply_reports_a_missing_target_schema(tmp_path, capsys):
    workspace = _workspace(tmp_path)
    (workspace / "schema.yaml").unlink()
    out = tmp_path / "merged.csv"

    exit_code = main(_apply_argv(workspace, out, "csv"))

    assert exit_code == 2
    assert not out.exists()
    assert "Hedef şema bulunamadı" in capsys.readouterr().out


def test_apply_never_builds_an_llm_client(tmp_path, monkeypatch):
    workspace = _workspace(tmp_path)
    out = tmp_path / "merged.csv"

    def fail(*args, **kwargs):  # pragma: no cover - only runs on a regression
        raise AssertionError("Phase 2 must not use an LLM")

    monkeypatch.setattr("cli.main.create_llm_client", fail)

    assert main(_apply_argv(workspace, out, "csv")) == 0
