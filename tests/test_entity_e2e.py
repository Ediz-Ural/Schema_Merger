"""Full flow: analyze -> mapping approval -> cluster -> cluster approval -> apply.

Phase 1 (``analyze``, ``cluster``) uses fake providers, phase 2 (``apply``) uses
none at all, so no test here makes a network request.
"""

from __future__ import annotations

import dataclasses
import json
import math
from pathlib import Path

from openpyxl import load_workbook

from cli.main import REVIEW_GUARD_EXIT_CODE, main
from core.contracts import ClusterContract, MappingContract, dump_clusters, dump_mapping, load_clusters, load_mapping
from core.llm import FakeEmbeddingClient, FakeLLMClient
from core.transformer import ENTITY_CLUSTER_COLUMN, MERGED_ROW_COUNT_COLUMN, original_value_column


SCHEMA_YAML = """target_columns:
  - name: product_name
    type: string
    required: true
  - name: unit_price
    type: decimal
    required: true
  - name: sold_on
    type: date
    required: true
output:
  format: xlsx
  add_provenance: true
"""

#: One vector per comparison key.  "coca cola 330ml" is what all three Cola
#: spellings normalise to, so they score 1.0 against each other.
VECTORS = {
    "coca cola 330ml": [1.0, 0.0],
    "fanta 330ml": [0.0, 1.0],
    # 0.72 against "fanta 330ml": inside the grey band, so it needs a human.
    "fanta portakal 330ml": [math.sqrt(1.0 - 0.72**2), 0.72],
}


def _mapping_response(product: str, price: str, sold_on: str) -> str:
    return json.dumps(
        {
            "matches": [
                {"target_column": "product_name", "column": product, "confidence": 0.95, "reason": "Adlar örtüşüyor."},
                {"target_column": "unit_price", "column": price, "confidence": 0.94, "reason": "Ondalık örnekler örtüşüyor."},
                {"target_column": "sold_on", "column": sold_on, "confidence": 0.93, "reason": "İkisi de tarih."},
            ]
        }
    )


def _approve_mapping(mapping_path: Path) -> None:
    mapping = load_mapping(mapping_path)
    entries = [
        dataclasses.replace(
            entry,
            sources=[
                dataclasses.replace(source, status="auto", confidence=1.0)
                if source.status == "review"
                else source
                for source in entry.sources
            ],
        )
        for entry in mapping.entries
    ]
    dump_mapping(MappingContract(entries=entries), mapping_path)


def _write_sources(tmp_path: Path) -> tuple[Path, Path, Path]:
    sales = tmp_path / "sales_2023.csv"
    sales.write_text(
        "urun;fiyat;tarih\n"
        "Coca Cola 330ml;12,50;31.01.2024\n"
        "Coca-Cola 33cl;12,50;31.01.2024\n"
        "Coca Cola 330ml;9,90;05.02.2024\n"
        "Fanta 330ml;10,00;01.02.2024\n",
        encoding="utf-8",
    )
    export = tmp_path / "export_q4.csv"
    export.write_text(
        "product,price,date\n"
        "coca cola 0.33 lt,8.90,2024-03-15\n"
        "Fanta Portakal 330ml,10.00,2024-02-01\n",
        encoding="utf-8",
    )
    schema = tmp_path / "schema.yaml"
    schema.write_text(SCHEMA_YAML, encoding="utf-8")
    return sales, export, schema


def _analyze_and_approve(tmp_path: Path, sales: Path, export: Path, schema: Path) -> Path:
    mapping_path = tmp_path / "mapping.yaml"
    client = FakeLLMClient(
        responses=[
            _mapping_response("urun", "fiyat", "tarih"),
            _mapping_response("product", "price", "date"),
        ]
    )
    assert (
        main(
            [
                "analyze",
                "--inputs",
                str(sales),
                str(export),
                "--target-schema",
                str(schema),
                "--out",
                str(mapping_path),
            ],
            llm_client=client,
        )
        == 0
    )
    _approve_mapping(mapping_path)
    return mapping_path


def test_cluster_proposes_groups_and_apply_deduplicates_only_approved_ones(tmp_path, capsys):
    sales, export, schema = _write_sources(tmp_path)
    mapping_path = _analyze_and_approve(tmp_path, sales, export, schema)
    clusters_path = tmp_path / "clusters.yaml"
    merged = tmp_path / "merged.xlsx"
    capsys.readouterr()

    # --- Phase 1b: cluster (embeddings + LLM for the grey band only) ----------
    embedder = FakeEmbeddingClient(vectors=VECTORS)
    grey_llm = FakeLLMClient(
        responses=[json.dumps({"same": True, "confidence": 0.7, "reason": "Aynı ürün olabilir."})]
    )
    code = main(
        [
            "cluster",
            "--mapping",
            str(mapping_path),
            "--column",
            "product_name",
            "--out",
            str(clusters_path),
        ],
        llm_client=grey_llm,
        embedding_client=embedder,
    )
    assert code == 0
    output = capsys.readouterr().out
    assert "1 küme yüksek güvenle birleşmeye hazır" in output
    assert "1 küme onay bekliyor" in output
    # Only the grey pair reached the LLM, and only the compared names were sent.
    assert len(grey_llm.calls or []) == 1
    assert "Fanta Portakal 330ml" in (grey_llm.calls or [])[0][1]
    assert "12,50" not in (grey_llm.calls or [])[0][1]

    contract = load_clusters(clusters_path)
    cola, fanta = contract.clusters
    assert cola.status == "auto"
    assert sorted(member.value for member in cola.members) == [
        "Coca Cola 330ml",
        "Coca-Cola 33cl",
        "coca cola 0.33 lt",
    ]
    assert fanta.status == "review"
    assert [candidate.value for candidate in fanta.candidates] == ["Fanta Portakal 330ml"]

    # --- Phase 2: apply with the reviewed cluster file (no LLM) ---------------
    apply_code = main(
        [
            "apply",
            "--mapping",
            str(mapping_path),
            "--out",
            str(merged),
            "--format",
            "xlsx",
            "--clusters",
            str(clusters_path),
        ]
    )
    assert apply_code == 0
    assert len(grey_llm.calls or []) == 1  # apply made no LLM call

    sheet = load_workbook(merged)["merged"]
    header = [cell.value for cell in sheet[1]]
    rows = [dict(zip(header, (cell.value for cell in row))) for row in sheet.iter_rows(min_row=2)]

    # 6 source rows -> 5: the two 12,50 Cola rows were one product spelled twice.
    assert len(rows) == 5
    assert [row["product_name"] for row in rows] == [
        "Coca Cola 330ml",
        "Coca Cola 330ml",
        "Fanta 330ml",
        "Coca Cola 330ml",
        # The unapproved Fanta cluster kept its own spelling.
        "Fanta Portakal 330ml",
    ]
    merged_row = rows[0]
    assert merged_row[ENTITY_CLUSTER_COLUMN] == "c001"
    assert merged_row[MERGED_ROW_COUNT_COLUMN] == 2
    assert merged_row["_source_file"] == "sales_2023.csv"
    assert merged_row[original_value_column("product_name")] == "Coca Cola 330ml; Coca-Cola 33cl"
    # The row from the other file kept its own origin and was not collapsed.
    assert rows[3]["_source_file"] == "export_q4.csv"
    assert rows[3][original_value_column("product_name")] == "coca cola 0.33 lt"

    console = capsys.readouterr().out
    assert "5 satır yazıldı" in console
    assert "1 onaylı küme uygulandı, 1 yinelenen satır tekilleştirildi" in console
    assert "1 küme onaysız kaldı" in console

    report = load_workbook(tmp_path / "merge_report.xlsx")
    entity_rows = [
        dict(zip([cell.value for cell in report["Entity"][1]], (cell.value for cell in row)))
        for row in report["Entity"].iter_rows(min_row=2)
    ]
    assert [row["status"] for row in entity_rows] == ["auto", "review"]
    assert "BELİRSİZ" in entity_rows[1]["note"]
    assert entity_rows[1]["candidates"] == "Fanta Portakal 330ml"


def test_approving_a_candidate_merges_it_on_the_next_apply(tmp_path, capsys):
    sales, export, schema = _write_sources(tmp_path)
    mapping_path = _analyze_and_approve(tmp_path, sales, export, schema)
    clusters_path = tmp_path / "clusters.yaml"
    merged = tmp_path / "merged.xlsx"
    assert (
        main(
            [
                "cluster",
                "--mapping",
                str(mapping_path),
                "--column",
                "product_name",
                "--out",
                str(clusters_path),
                "--no-llm",
            ],
            embedding_client=FakeEmbeddingClient(vectors=VECTORS),
        )
        == 0
    )

    # The user accepts the candidate: it becomes a member and the cluster auto.
    contract = load_clusters(clusters_path)
    cola, fanta = contract.clusters
    approved = dataclasses.replace(
        fanta,
        status="auto",
        members=[*fanta.members, dataclasses.replace(fanta.members[0], value="Fanta Portakal 330ml", row_count=1)],
        candidates=[],
    )
    dump_clusters(ClusterContract(clusters=[cola, approved]), clusters_path)
    capsys.readouterr()

    assert (
        main(
            [
                "apply",
                "--mapping",
                str(mapping_path),
                "--out",
                str(merged),
                "--format",
                "xlsx",
                "--clusters",
                str(clusters_path),
            ]
        )
        == 0
    )

    sheet = load_workbook(merged)["merged"]
    header = [cell.value for cell in sheet[1]]
    rows = [dict(zip(header, (cell.value for cell in row))) for row in sheet.iter_rows(min_row=2)]
    products = [row["product_name"] for row in rows]
    assert "Fanta Portakal 330ml" not in products  # it now says Fanta 330ml
    (fanta,) = [row for row in rows if row["product_name"] == "Fanta 330ml"]
    # The two Fanta rows describe the same sale of the same product, so they
    # collapsed and the surviving row still names both origins.
    assert fanta[MERGED_ROW_COUNT_COLUMN] == 2
    assert fanta["_source_file"] == "sales_2023.csv; export_q4.csv"
    assert fanta[original_value_column("product_name")] == "Fanta 330ml; Fanta Portakal 330ml"
    console = capsys.readouterr().out
    assert "2 onaylı küme uygulandı, 2 yinelenen satır tekilleştirildi" in console
    assert "onaysız kaldı" not in console


def test_cluster_refuses_a_column_that_is_not_text(tmp_path, capsys):
    sales, export, schema = _write_sources(tmp_path)
    mapping_path = _analyze_and_approve(tmp_path, sales, export, schema)
    capsys.readouterr()

    code = main(
        [
            "cluster",
            "--mapping",
            str(mapping_path),
            "--column",
            "unit_price",
            "--out",
            str(tmp_path / "clusters.yaml"),
        ],
        embedding_client=FakeEmbeddingClient(vectors=VECTORS),
    )

    assert code == 2
    assert "metin sütunlarında" in capsys.readouterr().out
    assert not (tmp_path / "clusters.yaml").exists()


def test_apply_accepts_an_empty_cluster_file_and_merges_nothing(tmp_path, capsys):
    sales, export, schema = _write_sources(tmp_path)
    mapping_path = _analyze_and_approve(tmp_path, sales, export, schema)
    clusters_path = tmp_path / "clusters.yaml"
    dump_clusters(ClusterContract(clusters=[]), clusters_path)
    merged = tmp_path / "merged.xlsx"
    capsys.readouterr()

    assert (
        main(
            ["apply", "--mapping", str(mapping_path), "--out", str(merged), "--clusters", str(clusters_path)]
        )
        == 0
    )

    sheet = load_workbook(merged)["merged"]
    assert sheet.max_row == 7  # header + the six untouched source rows
    report = load_workbook(tmp_path / "merge_report.xlsx")
    assert [cell.value for cell in report["Entity"][2]][5] == "empty"


def test_cluster_stops_while_the_mapping_still_waits_for_approval(tmp_path, capsys):
    sales, export, schema = _write_sources(tmp_path)
    mapping_path = tmp_path / "mapping.yaml"
    clusters_path = tmp_path / "clusters.yaml"
    client = FakeLLMClient(
        responses=[
            _mapping_response("urun", "fiyat", "tarih"),
            json.dumps(
                {
                    "matches": [
                        {"target_column": "product_name", "column": "product", "confidence": 0.55, "reason": "Belirsiz."},
                        {"target_column": "unit_price", "column": "price", "confidence": 0.94, "reason": "Ondalık."},
                        {"target_column": "sold_on", "column": "date", "confidence": 0.93, "reason": "Tarih."},
                    ]
                }
            ),
        ]
    )
    assert (
        main(
            ["analyze", "--inputs", str(sales), str(export), "--target-schema", str(schema), "--out", str(mapping_path)],
            llm_client=client,
        )
        == 0
    )
    capsys.readouterr()

    code = main(
        ["cluster", "--mapping", str(mapping_path), "--column", "product_name", "--out", str(clusters_path)],
        embedding_client=FakeEmbeddingClient(vectors=VECTORS),
    )

    assert code == REVIEW_GUARD_EXIT_CODE
    assert not clusters_path.exists()
    output = capsys.readouterr().out
    assert "cluster durdu" in output
    assert "product_name ← export_q4.csv:product" in output
