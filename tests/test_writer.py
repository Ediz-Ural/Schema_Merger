from __future__ import annotations

import sqlite3
from pathlib import Path

import polars as pl
import pytest
from openpyxl import load_workbook

from core.contracts import load_mapping, load_schema
from core.transformer import transform
from core.writer import (
    DEFAULT_REPORT_NAME,
    SOURCE_FILE_COLUMN,
    WriteError,
    source_column_name,
    write,
    write_merged,
    write_merge_report,
)


FIXTURES = Path(__file__).parent / "fixtures"
SOURCES = [FIXTURES / "transform_sales_tr.csv", FIXTURES / "transform_export_en.csv"]


@pytest.fixture()
def merge_context():
    mapping = load_mapping(FIXTURES / "transform_mapping.yaml")
    schema = load_schema(FIXTURES / "transform_schema.yaml")
    result = transform(mapping, SOURCES, schema, chunk_size=1)
    return result, mapping, schema


def _provenance_columns(schema) -> list[str]:
    return [SOURCE_FILE_COLUMN, *(source_column_name(c.name) for c in schema.target_columns)]


def test_csv_output_carries_data_and_provenance(merge_context, tmp_path):
    result, _, schema = merge_context
    out = tmp_path / "merged.csv"

    write_merged(result, schema, out, output_format="csv")

    frame = pl.read_csv(out)
    target_names = [c.name for c in schema.target_columns]
    assert frame.columns == [*target_names, *_provenance_columns(schema)]
    assert frame["product_name"].to_list() == ["Kalem", "Defter", "Pencil", "Notebook"]
    assert frame[SOURCE_FILE_COLUMN].to_list() == [
        "transform_sales_tr.csv",
        "transform_sales_tr.csv",
        "transform_export_en.csv",
        "transform_export_en.csv",
    ]
    assert frame[source_column_name("unit_price")].to_list() == ["fiyat", "fiyat", "price", "price"]
    # Unmatched target -> provenance column is empty for every row.
    assert frame[source_column_name("stock_quantity")].null_count() == 4


def test_xlsx_output_matches_the_merged_table(merge_context, tmp_path):
    result, _, schema = merge_context
    out = tmp_path / "merged.xlsx"

    write_merged(result, schema, out, output_format="xlsx")

    workbook = load_workbook(out)
    sheet = workbook["merged"]
    header = [cell.value for cell in sheet[1]]
    target_names = [c.name for c in schema.target_columns]
    assert header == [*target_names, *_provenance_columns(schema)]
    assert sheet.max_row == result.row_count + 1
    first_data_row = [cell.value for cell in sheet[2]]
    assert first_data_row[0] == "Kalem"
    assert first_data_row[header.index(SOURCE_FILE_COLUMN)] == "transform_sales_tr.csv"


def test_sql_output_is_a_runnable_create_and_insert_script(merge_context, tmp_path):
    result, _, schema = merge_context
    out = tmp_path / "merged.sql"

    write_merged(result, schema, out, output_format="sql", table_name="merged_products")

    script = out.read_text(encoding="utf-8")
    assert "CREATE TABLE" in script
    assert script.count("INSERT INTO") == result.row_count

    connection = sqlite3.connect(":memory:")
    try:
        connection.executescript(script)
        rows = connection.execute(
            'SELECT product_name, unit_price, "_source_file" FROM merged_products ORDER BY rowid'
        ).fetchall()
    finally:
        connection.close()
    assert [row[0] for row in rows] == ["Kalem", "Defter", "Pencil", "Notebook"]
    assert rows[0][1] == 12.5
    assert rows[0][2] == "transform_sales_tr.csv"


def test_merge_report_summarizes_matches_rows_and_nulls(merge_context, tmp_path):
    result, mapping, schema = merge_context
    report = tmp_path / "merge_report.xlsx"

    write_merge_report(result, mapping, schema, report, output_format="csv")

    workbook = load_workbook(report)
    assert {"Summary", "Columns", "Entity"}.issubset(set(workbook.sheetnames))

    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}
    assert summary["merged_rows"] == result.row_count
    assert summary["target_columns"] == len(schema.target_columns)
    assert summary["total_null_cells"] == 4  # stock_quantity is unmatched for every row
    assert summary["provenance_added"] is True

    columns = workbook["Columns"]
    header = [cell.value for cell in columns[1]]
    records = [dict(zip(header, (cell.value for cell in row))) for row in columns.iter_rows(min_row=2)]
    unit_price_rows = [r for r in records if r["target_column"] == "unit_price"]
    assert {r["source_column"] for r in unit_price_rows} == {"fiyat", "price"}
    assert all(r["status"] == "auto" for r in unit_price_rows)
    stock_rows = [r for r in records if r["target_column"] == "stock_quantity"]
    assert all(r["null_count"] == 4 and r["null_ratio"] == 1.0 for r in stock_rows)
    assert all(r["status"] == "unmatched" for r in stock_rows)


def test_write_emits_both_files_and_a_summary(merge_context, tmp_path):
    result, mapping, schema = merge_context
    out = tmp_path / "merged.csv"

    outcome = write(result, mapping, schema, out, output_format="csv")

    assert outcome.merged_path == out and out.exists()
    assert outcome.report_path == tmp_path / DEFAULT_REPORT_NAME and outcome.report_path.exists()
    assert outcome.output_format == "csv"
    assert outcome.row_count == result.row_count
    assert outcome.null_cell_count == 4
    assert outcome.conversion_error_count == 0


def test_provenance_is_written_even_when_flag_is_disabled(merge_context, tmp_path):
    result, _, schema = merge_context
    out = tmp_path / "merged.csv"

    write_merged(result, schema, out, output_format="csv")
    frame = pl.read_csv(out)

    # The invariant (spec §14/§6) is that provenance is always present.
    for column in _provenance_columns(schema):
        assert column in frame.columns


def test_unknown_format_is_rejected(merge_context, tmp_path):
    result, _, schema = merge_context
    with pytest.raises(WriteError):
        write_merged(result, schema, tmp_path / "merged.parquet", output_format="parquet")
