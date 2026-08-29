"""Edge cases the spec calls out (§11): multi-sheet workbooks, conflicting
values, type conflicts, and a missing API key.

These tests pin behaviour that is easy to regress silently: no row is invented,
no row is dropped, nothing is auto-corrected, and Phase 2 never needs a key.
"""

from __future__ import annotations

import json
from pathlib import Path

import polars as pl
import pytest
from openpyxl import Workbook, load_workbook

from cli.main import main
from core.contracts import load_mapping
from core.llm import FakeLLMClient, LLMConfigurationError, create_llm_client


SCHEMA_YAML = """target_columns:
  - name: product_name
    type: string
    required: true
  - name: unit_price
    type: decimal
    required: false
output:
  format: csv
  add_provenance: true
"""


def _schema(tmp_path: Path) -> Path:
    path = tmp_path / "schema.yaml"
    path.write_text(SCHEMA_YAML, encoding="utf-8")
    return path


def _mapping(tmp_path: Path, sources: list[tuple[str, str | None, str | None]]) -> Path:
    """Write a mapping plan where every listed source is already approved."""

    def block(column: str | None) -> str:
        return f"      column: {'null' if column is None else column}\n      confidence: 0.99\n      status: auto\n      reason: test\n"

    product = "".join(f"    - file: {name}\n{block(product_column)}" for name, product_column, _ in sources)
    price = "".join(f"    - file: {name}\n{block(price_column)}" for name, _, price_column in sources)
    path = tmp_path / "mapping.yaml"
    path.write_text(
        "- target_column: product_name\n  sources:\n" + product
        + "- target_column: unit_price\n  sources:\n" + price,
        encoding="utf-8",
    )
    return path


def _workbook(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    return path


def _apply(tmp_path: Path, *extra: str) -> tuple[int, Path]:
    out = tmp_path / "merged.csv"
    exit_code = main(
        ["apply", "--mapping", str(tmp_path / "mapping.yaml"), "--out", str(out), "--format", "csv", *extra]
    )
    return exit_code, out


# --------------------------------------------------------------------------
# Multi-sheet workbooks
# --------------------------------------------------------------------------


def test_every_worksheet_is_appended_and_named_in_provenance(tmp_path, capsys):
    _schema(tmp_path)
    _workbook(
        tmp_path / "book.xlsx",
        {
            "Ocak": [["urun", "fiyat"], ["Kalem", "12,50"]],
            "Subat": [["urun", "fiyat"], ["Defter", "8,00"]],
        },
    )
    _mapping(tmp_path, [("book.xlsx", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame["product_name"].to_list() == ["Kalem", "Defter"]
    # The sheet a row came from stays visible in provenance (spec §6).
    assert frame["_source_file"].to_list() == ["book.xlsx#Ocak", "book.xlsx#Subat"]
    assert "atlanan sheet" not in capsys.readouterr().out


def test_sheet_flag_restricts_the_run_to_one_worksheet(tmp_path):
    _schema(tmp_path)
    _workbook(
        tmp_path / "book.xlsx",
        {
            "Ocak": [["urun", "fiyat"], ["Kalem", "12,50"]],
            "Subat": [["urun", "fiyat"], ["Defter", "8,00"]],
        },
    )
    _mapping(tmp_path, [("book.xlsx", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path, "--sheet", "Subat")

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame["product_name"].to_list() == ["Defter"]
    # A single worksheet needs no qualifier to be unambiguous.
    assert frame["_source_file"].to_list() == ["book.xlsx"]


def test_unknown_sheet_name_is_reported_with_the_available_ones(tmp_path, capsys):
    _schema(tmp_path)
    _workbook(tmp_path / "book.xlsx", {"Ocak": [["urun", "fiyat"], ["Kalem", "12,50"]]})
    _mapping(tmp_path, [("book.xlsx", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path, "--sheet", "Mart")

    assert exit_code == 2
    assert not out.exists()
    output = capsys.readouterr().out
    assert "'Mart' sheet'i yok" in output
    assert "Ocak" in output


def test_unrelated_worksheet_is_skipped_instead_of_adding_empty_rows(tmp_path, capsys):
    _schema(tmp_path)
    _workbook(
        tmp_path / "book.xlsx",
        {
            "Satis": [["urun", "fiyat"], ["Kalem", "12,50"]],
            "Adresler": [["sehir", "telefon"], ["Ankara", "0312"]],
        },
    )
    _mapping(tmp_path, [("book.xlsx", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame.height == 1  # the address sheet contributes no empty row
    output = capsys.readouterr().out
    assert "atlanan sheet: book.xlsx#Adresler" in output
    summary = {
        row[0].value: row[1].value
        for row in load_workbook(tmp_path / "merge_report.xlsx")["Summary"].iter_rows(min_row=2)
    }
    assert summary["skipped_sheets"] == "book.xlsx#Adresler"


def test_analyze_sheet_flag_profiles_only_that_worksheet(tmp_path):
    _workbook(
        tmp_path / "book.xlsx",
        {
            "Satis": [["urun", "fiyat"], ["Kalem", "12,50"]],
            "Adresler": [["sehir", "telefon"], ["Ankara", "0312"]],
        },
    )
    client = FakeLLMClient(
        response=json.dumps(
            {
                "matches": [
                    {"target_column": "product_name", "column": "urun", "confidence": 0.95, "reason": "ok"},
                    {"target_column": "unit_price", "column": "fiyat", "confidence": 0.95, "reason": "ok"},
                ]
            }
        )
    )

    exit_code = main(
        [
            "analyze",
            "--inputs",
            str(tmp_path / "book.xlsx"),
            "--target-schema",
            str(_schema(tmp_path)),
            "--out",
            str(tmp_path / "mapping.yaml"),
            "--sheet",
            "Satis",
        ],
        llm_client=client,
    )

    assert exit_code == 0
    prompt = (client.calls or [("", "")])[0][1]
    assert "urun" in prompt and "sehir" not in prompt
    assert [entry.target_column for entry in load_mapping(tmp_path / "mapping.yaml").entries] == [
        "product_name",
        "unit_price",
    ]


# --------------------------------------------------------------------------
# Conflicting values
# --------------------------------------------------------------------------


def test_conflicting_values_are_all_kept_and_marked_by_provenance(tmp_path):
    """Two files disagree on one product's price; neither is picked (spec §11)."""

    _schema(tmp_path)
    (tmp_path / "kasa.csv").write_text("urun;fiyat\nKalem;12,50\n", encoding="utf-8")
    (tmp_path / "depo.csv").write_text("product,price\nKalem,19.90\n", encoding="utf-8")
    _mapping(tmp_path, [("kasa.csv", "urun", "fiyat"), ("depo.csv", "product", "price")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    frame = pl.read_csv(out)
    # Both readings survive; no row is dropped and no value is "corrected".
    assert frame["product_name"].to_list() == ["Kalem", "Kalem"]
    assert sorted(frame["unit_price"].to_list()) == [12.5, 19.9]
    assert frame["_source_file"].to_list() == ["kasa.csv", "depo.csv"]
    assert frame["unit_price_source_column"].to_list() == ["fiyat", "price"]


# --------------------------------------------------------------------------
# Type conflicts
# --------------------------------------------------------------------------


def test_unconvertible_value_is_nulled_counted_and_reported_not_dropped(tmp_path, capsys):
    """One broken cell among many: the row survives, the failure is counted.

    The ratio stays under the validator's type threshold, so this is a report,
    not a stop — see the next test for the case where the column looks wrong.
    """

    _schema(tmp_path)
    rows = "\n".join(f"Kalem{index};1{index},50" for index in range(9))
    (tmp_path / "kasa.csv").write_text(f"urun;fiyat\n{rows}\nDefter;fiyat yok\n", encoding="utf-8")
    _mapping(tmp_path, [("kasa.csv", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    frame = pl.read_csv(out)
    assert frame.height == 10  # the broken row stays, only its cell is null
    assert frame["unit_price"].to_list()[-1] is None
    assert frame["product_name"].to_list()[-1] == "Defter"

    output = capsys.readouterr().out
    assert "1 dönüştürme hatası" in output

    report = load_workbook(tmp_path / "merge_report.xlsx")
    summary = {row[0].value: row[1].value for row in report["Summary"].iter_rows(min_row=2)}
    assert summary["total_conversion_errors"] == 1
    columns = list(report["Columns"].iter_rows(min_row=2, values_only=True))
    header = [cell.value for cell in report["Columns"][1]]
    price_row = dict(zip(header, next(row for row in columns if row[0] == "unit_price")))
    assert price_row["conversion_errors"] == 1


def test_a_column_of_the_wrong_type_stops_the_merge_instead_of_merging_blind(tmp_path, capsys):
    _schema(tmp_path)
    (tmp_path / "kasa.csv").write_text(
        "urun;fiyat\nKalem;12,50\nDefter;fiyat yok\nSilgi;3,25\n", encoding="utf-8"
    )
    _mapping(tmp_path, [("kasa.csv", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 3
    assert not out.exists()  # nothing is written, nothing is silently discarded
    output = capsys.readouterr().out
    assert "apply durdu: validator" in output
    assert "unit_price ← kasa.csv:fiyat" in output


def test_turkish_and_english_numbers_normalize_side_by_side(tmp_path):
    _schema(tmp_path)
    (tmp_path / "kasa.csv").write_text("urun;fiyat\nKalem;1.234,56\n", encoding="utf-8")
    (tmp_path / "depo.csv").write_text("product,price\nPencil,1234.56\n", encoding="utf-8")
    _mapping(tmp_path, [("kasa.csv", "urun", "fiyat"), ("depo.csv", "product", "price")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    assert pl.read_csv(out)["unit_price"].to_list() == [1234.56, 1234.56]


# --------------------------------------------------------------------------
# Missing API key
# --------------------------------------------------------------------------


@pytest.fixture
def no_env_keys(monkeypatch):
    """An environment with no provider configuration and no .env fallback."""

    monkeypatch.setattr("core.llm.load_dotenv", lambda *args, **kwargs: False)
    for variable in ("LLM_PROVIDER", "EMBEDDING_PROVIDER", "OPENAI_API_KEY", "ANTHROPIC_API_KEY"):
        monkeypatch.delenv(variable, raising=False)


def test_analyze_without_a_key_fails_with_a_clear_message(tmp_path, capsys, no_env_keys):
    (tmp_path / "kasa.csv").write_text("urun;fiyat\nKalem;12,50\n", encoding="utf-8")
    output = tmp_path / "mapping.yaml"

    exit_code = main(
        [
            "analyze",
            "--inputs",
            str(tmp_path / "kasa.csv"),
            "--target-schema",
            str(_schema(tmp_path)),
            "--out",
            str(output),
        ]
    )

    assert exit_code == 2
    assert not output.exists()
    message = capsys.readouterr().out
    assert "OPENAI_API_KEY tanımlı değil" in message
    assert ".env" in message


def test_the_key_error_names_the_variable_of_the_selected_provider(no_env_keys, monkeypatch):
    monkeypatch.setenv("LLM_PROVIDER", "anthropic")

    with pytest.raises(LLMConfigurationError, match="ANTHROPIC_API_KEY"):
        create_llm_client()


def test_apply_needs_no_key_at_all(tmp_path, no_env_keys):
    _schema(tmp_path)
    (tmp_path / "kasa.csv").write_text("urun;fiyat\nKalem;12,50\n", encoding="utf-8")
    _mapping(tmp_path, [("kasa.csv", "urun", "fiyat")])

    exit_code, out = _apply(tmp_path)

    assert exit_code == 0
    assert pl.read_csv(out).height == 1


# --------------------------------------------------------------------------
# README smoke: the documented commands really run
# --------------------------------------------------------------------------


FIXTURES = Path(__file__).parent / "fixtures"


@pytest.mark.parametrize(
    "argv",
    [
        ["profile", "--input", str(FIXTURES / "sample_tr.csv")],
        ["profile", "--input", str(FIXTURES / "sample_multi_sheet.xlsx"), "--sheet", "Stok"],
        ["profile", "--input", str(FIXTURES / "sample_multi_sheet.xlsx")],
    ],
)
def test_readme_profile_examples_run(argv, capsys):
    assert main(argv) == 0
    assert "Table:" in capsys.readouterr().out
