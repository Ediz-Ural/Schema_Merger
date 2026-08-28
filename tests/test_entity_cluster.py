"""Layer 5 of entity resolution: clusters, cluster approval, and deduplication.

Every embedder and LLM here is a fake, so no test makes a network request.
"""

from __future__ import annotations

import dataclasses
import math

import polars as pl
import pytest
from openpyxl import load_workbook

from core.contracts import (
    ClusterContract,
    ContractValidationError,
    MappingContract,
    OutputSettings,
    SchemaContract,
    canonical_map,
    dump_clusters,
    load_clusters,
    pending_clusters,
)
from core.entity import candidate_pairs, cluster, make_blocks, score_pairs, to_cluster_plans
from core.llm import FakeEmbeddingClient
from core.transformer import (
    ENTITY_CLUSTER_COLUMN,
    MERGED_ROW_COUNT_COLUMN,
    TransformError,
    TransformationResult,
    deduplicate,
    original_value_column,
)
from core.types import (
    ClusterCandidate,
    ClusterMember,
    EntityClusterPlan,
    MappingEntry,
    SourceMatch,
    TargetColumn,
)
from core.writer import EntitySummary, write_merge_report


REFERENCE = [1.0, 0.0]

#: Three spellings of one product; all three normalise to "coca cola 330ml".
VARIANTS = ["Coca Cola 330ml", "Coca-Cola 33cl", "coca cola 0,33 lt"]


def at_similarity(value: float) -> list[float]:
    """A unit vector whose cosine similarity with :data:`REFERENCE` is ``value``."""

    return [value, math.sqrt(max(0.0, 1.0 - value * value))]


def decisions_for(records: list[dict[str, object]], vectors: dict[str, list[float]]) -> list:
    blocks = make_blocks(records)
    embedder = FakeEmbeddingClient(vectors=vectors)
    return score_pairs(candidate_pairs(blocks), embedder)


def variant_records(counts: tuple[int, int, int] = (5, 2, 1)) -> list[dict[str, object]]:
    return [{"name": name, "row_count": count} for name, count in zip(VARIANTS, counts)]


# --------------------------------------------------------------------------- #
# Clustering
# --------------------------------------------------------------------------- #


def test_variant_group_collapses_into_one_cluster_with_a_canonical_value() -> None:
    result = cluster(
        decisions_for(variant_records(), {"coca cola 330ml": REFERENCE}),
        target_column="product_name",
    )

    (group,) = result.clusters
    assert group.cluster_id == "c001"
    assert sorted(group.values) == sorted(VARIANTS)
    # Most merged rows wins, so the 5-row spelling represents the cluster.
    assert group.canonical_value == "Coca Cola 330ml"
    assert group.status == "auto"
    assert group.candidates == ()


def test_a_spelling_that_matches_nothing_is_not_reported_as_a_cluster() -> None:
    records = [{"name": "Coca Cola 330ml"}, {"name": "Coca Cola Zero 330ml"}]
    result = cluster(
        decisions_for(
            records,
            {"coca cola 330ml": REFERENCE, "coca cola zero 330ml": at_similarity(0.10)},
        )
    )

    # "different" needs no human decision, so nothing is put in front of one.
    assert result.clusters == ()


def test_a_grey_pair_becomes_a_candidate_and_keeps_the_cluster_in_review() -> None:
    records = variant_records() + [{"name": "Coca Cola Zero 330ml", "row_count": 4}]
    result = cluster(
        decisions_for(
            records,
            {"coca cola 330ml": REFERENCE, "coca cola zero 330ml": at_similarity(0.72)},
        ),
        target_column="product_name",
    )

    (group,) = result.clusters
    assert group.status == "review"
    # The three certain spellings are members; the uncertain one is only proposed.
    assert sorted(group.values) == sorted(VARIANTS)
    assert [link.record.value for link in group.candidates] == ["Coca Cola Zero 330ml"]
    assert group.candidates[0].decision.similarity == pytest.approx(0.72)
    assert "onay bekliyor" in group.reason


def test_the_same_candidate_seen_from_several_members_is_one_decision() -> None:
    records = variant_records() + [{"name": "Coca Cola Zero 330ml"}]
    result = cluster(
        decisions_for(
            records,
            {"coca cola 330ml": REFERENCE, "coca cola zero 330ml": at_similarity(0.72)},
        )
    )

    # Three grey pairs (one per member) but a single candidate to decide on.
    (group,) = result.clusters
    assert len(group.candidates) == 1


def test_high_confidence_clusters_are_auto_and_grey_ones_are_not() -> None:
    records = [
        {"name": "Coca Cola 330ml", "row_count": 3},
        {"name": "Coca-Cola 33cl", "row_count": 1},
        {"name": "Fanta 330ml", "row_count": 2},
        {"name": "Fanta Portakal 330ml", "row_count": 1},
    ]
    vectors = {
        "coca cola 330ml": REFERENCE,
        "fanta 330ml": [0.0, 1.0],
        "fanta portakal 330ml": at_similarity(0.70),
    }
    # Blocking keeps "coc" and "fan" apart, so only same-brand pairs are scored.
    result = cluster(decisions_for(records, vectors), target_column="product_name")

    statuses = {group.canonical_value: group.status for group in result.clusters}
    assert statuses["Coca Cola 330ml"] == "auto"
    assert statuses["Fanta 330ml"] == "review"
    assert [group.cluster_id for group in result.auto_clusters] == ["c001"]
    assert [group.cluster_id for group in result.review_clusters] == ["c002"]


# --------------------------------------------------------------------------- #
# clusters.yaml — the file the user edits
# --------------------------------------------------------------------------- #


def test_cluster_plans_round_trip_through_clusters_yaml(tmp_path) -> None:
    result = cluster(
        decisions_for(variant_records(), {"coca cola 330ml": REFERENCE}),
        target_column="product_name",
    )
    path = tmp_path / "clusters.yaml"
    dump_clusters(to_cluster_plans(result), path)

    text = path.read_text(encoding="utf-8")
    assert "status: auto" in text
    assert "# status: review" in text  # the approval rules travel with the file

    loaded = load_clusters(path)
    assert loaded.target_column == "product_name"
    assert [member.value for member in loaded.clusters[0].members] == VARIANTS
    assert loaded.clusters[0].canonical == "Coca Cola 330ml"
    assert [member.row_count for member in loaded.clusters[0].members] == [5, 2, 1]


def test_clusters_yaml_rejects_a_canonical_that_is_not_a_member(tmp_path) -> None:
    path = tmp_path / "clusters.yaml"
    path.write_text(
        "- cluster_id: c001\n"
        "  target_column: product_name\n"
        "  canonical: Pepsi\n"
        "  status: auto\n"
        "  members:\n"
        "    - value: Coca Cola 330ml\n",
        encoding="utf-8",
    )

    with pytest.raises(ContractValidationError, match="canonical"):
        load_clusters(path)


def test_clusters_yaml_rejects_a_value_shared_by_two_clusters(tmp_path) -> None:
    path = tmp_path / "clusters.yaml"
    path.write_text(
        "- cluster_id: c001\n"
        "  target_column: product_name\n"
        "  canonical: Coca Cola 330ml\n"
        "  status: auto\n"
        "  members:\n"
        "    - value: Coca Cola 330ml\n"
        "- cluster_id: c002\n"
        "  target_column: product_name\n"
        "  canonical: Coca Cola 330ml\n"
        "  status: auto\n"
        "  members:\n"
        "    - value: Coca Cola 330ml\n",
        encoding="utf-8",
    )

    with pytest.raises(ContractValidationError, match="tek kümeye ait"):
        load_clusters(path)


def test_clusters_yaml_rejects_an_unknown_status(tmp_path) -> None:
    path = tmp_path / "clusters.yaml"
    path.write_text(
        "- cluster_id: c001\n"
        "  target_column: product_name\n"
        "  canonical: Coca Cola 330ml\n"
        "  status: merged\n"
        "  members:\n"
        "    - value: Coca Cola 330ml\n",
        encoding="utf-8",
    )

    with pytest.raises(ContractValidationError, match="status"):
        load_clusters(path)


def test_canonical_map_ignores_everything_the_user_has_not_approved() -> None:
    approved = EntityClusterPlan(
        cluster_id="c001",
        target_column="product_name",
        canonical="Coca Cola 330ml",
        status="auto",
        members=[ClusterMember(value="Coca Cola 330ml"), ClusterMember(value="Coca-Cola 33cl")],
    )
    waiting = dataclasses.replace(
        approved,
        cluster_id="c002",
        status="review",
        canonical="Fanta 330ml",
        members=[ClusterMember(value="Fanta 330ml"), ClusterMember(value="Fanta Portakal 330ml")],
    )
    rejected = dataclasses.replace(
        approved,
        cluster_id="c003",
        status="rejected",
        canonical="Pepsi 330ml",
        members=[ClusterMember(value="Pepsi 330ml"), ClusterMember(value="Pepsi Max 330ml")],
    )

    mapping = canonical_map([approved, waiting, rejected])
    assert set(mapping) == {"Coca Cola 330ml", "Coca-Cola 33cl"}
    assert mapping["Coca-Cola 33cl"] == ("Coca Cola 330ml", "c001")
    assert [plan.cluster_id for plan in pending_clusters([approved, waiting, rejected])] == ["c002"]


# --------------------------------------------------------------------------- #
# Deduplication in the transformer
# --------------------------------------------------------------------------- #


SCHEMA = SchemaContract(
    target_columns=[
        TargetColumn(name="product_name", type="string", required=True),
        TargetColumn(name="unit_price", type="decimal", required=True),
    ],
    output=OutputSettings(format="xlsx", add_provenance=True),
)

MAPPING = MappingContract(
    entries=[
        MappingEntry(
            target_column="product_name",
            sources=[SourceMatch(file="sales.csv", column="urun", confidence=1.0, status="auto")],
        ),
        MappingEntry(
            target_column="unit_price",
            sources=[SourceMatch(file="sales.csv", column="fiyat", confidence=1.0, status="auto")],
        ),
    ]
)


def transformation(rows: list[tuple[str, float, str, str]]) -> TransformationResult:
    """Build a transformed table directly: (product, price, file, source column)."""

    return TransformationResult(
        dataframe=pl.DataFrame(
            {
                "product_name": [row[0] for row in rows],
                "unit_price": [row[1] for row in rows],
            },
            schema={"product_name": pl.String, "unit_price": pl.Float64},
        ),
        provenance=pl.DataFrame(
            {
                "source_file": [row[2] for row in rows],
                "product_name": [row[3] for row in rows],
                "unit_price": ["fiyat" for _ in rows],
            },
            schema={"source_file": pl.String, "product_name": pl.String, "unit_price": pl.String},
        ),
        conversion_error_counts={"product_name": 0, "unit_price": 0},
    )


APPROVED = EntityClusterPlan(
    cluster_id="c001",
    target_column="product_name",
    canonical="Coca Cola 330ml",
    status="auto",
    members=[
        ClusterMember(value="Coca Cola 330ml", row_count=1),
        ClusterMember(value="Coca-Cola 33cl", row_count=1),
    ],
)


def test_approved_clusters_canonicalize_values_and_collapse_duplicate_rows() -> None:
    result = transformation(
        [
            ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
            ("Coca-Cola 33cl", 12.5, "export.csv", "product"),
            ("Coca-Cola 33cl", 15.0, "export.csv", "product"),
            ("Fanta 330ml", 10.0, "sales.csv", "urun"),
        ]
    )

    dedup = deduplicate(result, [APPROVED])

    frame = dedup.result.dataframe
    assert frame.height == 3  # the two 12.50 rows were one product all along
    assert frame["product_name"].to_list() == ["Coca Cola 330ml", "Coca Cola 330ml", "Fanta 330ml"]
    assert frame["unit_price"].to_list() == [12.5, 15.0, 10.0]
    assert dedup.duplicate_row_count == 1
    assert dedup.canonicalized_row_count == 2
    assert dedup.merged_cluster_count == 1


def test_deduplication_keeps_the_origin_of_every_row_it_absorbs() -> None:
    result = transformation(
        [
            ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
            ("Coca-Cola 33cl", 12.5, "export.csv", "product"),
        ]
    )

    dedup = deduplicate(result, [APPROVED])

    (row,) = dedup.result.provenance.rows(named=True)
    assert row["source_file"] == "sales.csv; export.csv"
    assert row["product_name"] == "urun; product"
    assert row[ENTITY_CLUSTER_COLUMN] == "c001"
    assert row[original_value_column("product_name")] == "Coca Cola 330ml; Coca-Cola 33cl"
    assert row[MERGED_ROW_COUNT_COLUMN] == 2


def test_an_unapproved_cluster_never_merges_a_row() -> None:
    waiting = dataclasses.replace(APPROVED, status="review")
    rows = [
        ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
        ("Coca-Cola 33cl", 12.5, "export.csv", "product"),
    ]

    dedup = deduplicate(transformation(rows), [waiting])

    frame = dedup.result.dataframe
    assert frame.height == 2
    assert frame["product_name"].to_list() == ["Coca Cola 330ml", "Coca-Cola 33cl"]
    assert dedup.duplicate_row_count == 0
    assert dedup.merged_cluster_count == 0
    # Provenance still says which cluster touched the row: none of them.
    assert dedup.result.provenance[ENTITY_CLUSTER_COLUMN].to_list() == [None, None]


def test_identical_rows_outside_a_cluster_are_left_alone() -> None:
    rows = [
        ("Fanta 330ml", 10.0, "sales.csv", "urun"),
        ("Fanta 330ml", 10.0, "sales.csv", "urun"),
    ]

    dedup = deduplicate(transformation(rows), [APPROVED])

    # Blind deduplication is not this module's job; only clusters merge rows.
    assert dedup.result.dataframe.height == 2
    assert dedup.duplicate_row_count == 0


def test_a_spelling_repeated_inside_a_cluster_keeps_its_rows() -> None:
    rows = [
        ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
        ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
        ("Coca-Cola 33cl", 12.5, "export.csv", "product"),
    ]

    dedup = deduplicate(transformation(rows), [APPROVED])

    # The same product sold twice at the same price is two sales; only the
    # duplicate that entity resolution itself created is removed.
    assert dedup.result.dataframe.height == 2
    assert dedup.duplicate_row_count == 1
    assert dedup.result.provenance[MERGED_ROW_COUNT_COLUMN].to_list() == [2, 1]


def test_deduplication_rejects_a_column_that_is_not_in_the_merged_table() -> None:
    plan = dataclasses.replace(APPROVED, target_column="urun_adi")

    with pytest.raises(TransformError, match="urun_adi"):
        deduplicate(transformation([("Coca Cola 330ml", 12.5, "sales.csv", "urun")]), [plan])


def test_deduplication_refuses_a_non_text_entity_column() -> None:
    with pytest.raises(TransformError, match="metin değil"):
        deduplicate(
            transformation([("Coca Cola 330ml", 12.5, "sales.csv", "urun")]),
            [APPROVED],
            column="unit_price",
        )


# --------------------------------------------------------------------------- #
# merge_report
# --------------------------------------------------------------------------- #


def test_merge_report_lists_merged_and_uncertain_clusters(tmp_path) -> None:
    waiting = EntityClusterPlan(
        cluster_id="c002",
        target_column="product_name",
        canonical="Fanta 330ml",
        status="review",
        members=[ClusterMember(value="Fanta 330ml", row_count=2)],
        candidates=[
            ClusterCandidate(
                value="Fanta Portakal 330ml",
                similarity=0.71,
                suggestion="undecided",
                source="llm",
                reason="Gri bölge; karar kullanıcıda.",
            )
        ],
        reason="1 yazım kümede; 1 belirsiz aday onay bekliyor. Onaylanmadan birleştirilmez.",
    )
    result = transformation(
        [
            ("Coca Cola 330ml", 12.5, "sales.csv", "urun"),
            ("Coca-Cola 33cl", 12.5, "export.csv", "product"),
            ("Fanta 330ml", 10.0, "sales.csv", "urun"),
        ]
    )
    contract = ClusterContract(clusters=[APPROVED, waiting])
    dedup = deduplicate(result, contract)
    report = tmp_path / "merge_report.xlsx"

    write_merge_report(
        dedup.result,
        MAPPING,
        SCHEMA,
        report,
        entity=EntitySummary.from_deduplication(dedup, contract),
    )

    workbook = load_workbook(report)
    entity_rows = [
        dict(zip([cell.value for cell in workbook["Entity"][1]], (cell.value for cell in row)))
        for row in workbook["Entity"].iter_rows(min_row=2)
    ]
    assert [row["cluster_id"] for row in entity_rows] == ["c001", "c002"]
    assert entity_rows[0]["status"] == "auto"
    assert entity_rows[0]["members"] == "Coca Cola 330ml; Coca-Cola 33cl"
    assert entity_rows[1]["status"] == "review"
    assert "BELİRSİZ" in entity_rows[1]["note"]

    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_row=2)}
    assert summary["entity_column"] == "product_name"
    assert summary["entity_clusters_merged"] == 1
    assert summary["entity_pending_clusters"] == 1
    assert summary["entity_duplicate_rows_removed"] == 1
