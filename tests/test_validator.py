from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl
import pytest
from openpyxl import load_workbook

from cli.main import REVIEW_GUARD_EXIT_CODE, main
from core.contracts import MappingContract, OutputSettings, SchemaContract
from core.transformer import TransformationResult
from core.types import MappingEntry, SourceMatch, TargetColumn
from core.validator import (
    ValidationSettings,
    downgrade_to_review,
    validate,
)


def _schema(*columns: tuple[str, str, bool]) -> SchemaContract:
    return SchemaContract(
        target_columns=[
            TargetColumn(name=name, type=column_type, required=required)
            for name, column_type, required in columns
        ],
        output=OutputSettings(format="csv", add_provenance=True),
    )


def _mapping(**targets: list[tuple[str, str | None, str]]) -> MappingContract:
    """``target=[(file, column, status), ...]`` written as a mapping plan."""

    return MappingContract(
        entries=[
            MappingEntry(
                target_column=target,
                sources=[
                    SourceMatch(file=file, column=column, confidence=0.9, status=status)
                    for file, column, status in sources
                ],
            )
            for target, sources in targets.items()
        ]
    )


def _result(
    data: dict[str, list[object]],
    provenance: dict[str, list[object]],
    *,
    errors: dict[str, int] | None = None,
    dtypes: dict[str, pl.DataType] | None = None,
) -> TransformationResult:
    frame = pl.DataFrame(data)
    if dtypes:
        frame = frame.cast(dtypes, strict=False)
    return TransformationResult(
        dataframe=frame,
        provenance=pl.DataFrame(provenance).cast(
            {name: pl.String for name in provenance}, strict=False
        ),
        conversion_error_counts=errors or {name: 0 for name in data},
    )


# --------------------------------------------------------------------------- #
# Type consistency
# --------------------------------------------------------------------------- #


def test_many_conversion_failures_are_an_error():
    """Text in a decimal column: the matcher most likely picked the wrong one."""

    result = _result(
        {"unit_price": [12.5, None, None, None]},
        {"source_file": ["a.csv"] * 4, "unit_price": ["fiyat"] * 4},
        errors={"unit_price": 3},
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("unit_price", "decimal", False))
    mapping = _mapping(unit_price=[("a.csv", "fiyat", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "type")
    assert finding.severity == "error"
    assert finding.count == 3
    assert finding.source_file == "a.csv"
    assert finding.source_column == "fiyat"
    assert report.blocking


def test_a_few_conversion_failures_are_only_a_warning():
    result = _result(
        {"unit_price": [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, None]},
        {"source_file": ["a.csv"] * 10, "unit_price": ["fiyat"] * 10},
        errors={"unit_price": 1},
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("unit_price", "decimal", False))
    mapping = _mapping(unit_price=[("a.csv", "fiyat", "auto")])

    report = validate(result, mapping, schema)

    assert [finding.severity for finding in report.findings if finding.check == "type"] == ["warning"]
    assert not report.blocking


def test_a_dtype_that_disagrees_with_the_target_type_is_an_error():
    result = _result(
        {"stock_quantity": ["10", "3"]},
        {"source_file": ["a.csv"] * 2, "stock_quantity": ["stok"] * 2},
    )
    schema = _schema(("stock_quantity", "integer", False))
    mapping = _mapping(stock_quantity=[("a.csv", "stok", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "type")
    assert finding.severity == "error"
    assert "integer" in finding.message


# --------------------------------------------------------------------------- #
# Null explosion
# --------------------------------------------------------------------------- #


def test_null_explosion_is_flagged_for_the_mapped_source():
    result = _result(
        {"stock_quantity": [1, None, None, None]},
        {"source_file": ["a.csv"] * 4, "stock_quantity": ["stok"] * 4},
        dtypes={"stock_quantity": pl.Int64},
    )
    schema = _schema(("stock_quantity", "integer", False))
    mapping = _mapping(stock_quantity=[("a.csv", "stok", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "null")
    assert finding.severity == "warning"
    assert finding.metric == 0.75
    assert finding.source_file == "a.csv"


def test_a_mapped_source_that_is_entirely_null_is_an_error():
    result = _result(
        {"stock_quantity": [None, None, None]},
        {"source_file": ["a.csv"] * 3, "stock_quantity": ["stok"] * 3},
        dtypes={"stock_quantity": pl.Int64},
    )
    schema = _schema(("stock_quantity", "integer", False))
    mapping = _mapping(stock_quantity=[("a.csv", "stok", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "null")
    assert finding.severity == "error"
    assert "tamamen boş" in finding.message


def test_planned_nulls_from_an_unmapped_file_are_not_flagged():
    """Rows from a file that never mapped the column are not a defect."""

    result = _result(
        {"stock_quantity": [1, 2, None, None, None, None]},
        {
            "source_file": ["a.csv", "a.csv", "b.csv", "b.csv", "b.csv", "b.csv"],
            "stock_quantity": ["stok", "stok", None, None, None, None],
        },
        dtypes={"stock_quantity": pl.Int64},
    )
    schema = _schema(("stock_quantity", "integer", False))
    mapping = _mapping(stock_quantity=[("a.csv", "stok", "auto"), ("b.csv", None, "unmatched")])

    report = validate(result, mapping, schema)

    assert [item for item in report.findings if item.check == "null"] == []


def test_the_null_threshold_is_configurable():
    result = _result(
        {"stock_quantity": [1, 2, 3, None]},
        {"source_file": ["a.csv"] * 4, "stock_quantity": ["stok"] * 4},
        dtypes={"stock_quantity": pl.Int64},
    )
    schema = _schema(("stock_quantity", "integer", False))
    mapping = _mapping(stock_quantity=[("a.csv", "stok", "auto")])

    assert not [item for item in validate(result, mapping, schema).findings if item.check == "null"]

    strict = validate(result, mapping, schema, settings=ValidationSettings(null_warning_ratio=0.1))
    assert [item.severity for item in strict.findings if item.check == "null"] == ["warning"]


# --------------------------------------------------------------------------- #
# Outliers and formats
# --------------------------------------------------------------------------- #


def test_a_numeric_outlier_is_flagged_as_a_warning():
    values = [10.0, 11.0, 12.0, 10.5, 11.5, 9.5, 10.2, 11.1, 999_999.0]
    result = _result(
        {"unit_price": values},
        {"source_file": ["a.csv"] * len(values), "unit_price": ["fiyat"] * len(values)},
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("unit_price", "decimal", False))
    mapping = _mapping(unit_price=[("a.csv", "fiyat", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "outlier")
    assert finding.severity == "warning"
    assert finding.count == 1
    assert "999999" in finding.message
    assert not report.blocking  # an outlier is for a human to look at, not a block


def test_a_uniform_numeric_column_produces_no_outlier():
    values = [10.0] * 12
    result = _result(
        {"unit_price": values},
        {"source_file": ["a.csv"] * len(values), "unit_price": ["fiyat"] * len(values)},
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("unit_price", "decimal", False))
    mapping = _mapping(unit_price=[("a.csv", "fiyat", "auto")])

    assert [item for item in validate(result, mapping, schema).findings if item.check == "outlier"] == []


def test_an_implausible_date_is_flagged():
    dates = [date(2023, 5, 1), date(2023, 6, 1), date(1899, 1, 1)]
    result = _result(
        {"sold_at": dates},
        {"source_file": ["a.csv"] * 3, "sold_at": ["tarih"] * 3},
        dtypes={"sold_at": pl.Date},
    )
    schema = _schema(("sold_at", "date", False))
    mapping = _mapping(sold_at=[("a.csv", "tarih", "auto")])

    finding = next(item for item in validate(result, mapping, schema).findings if item.check == "outlier")
    assert finding.severity == "warning"
    assert "1899-01-01" in finding.message


def test_mixed_decimal_formats_in_a_text_column_are_flagged():
    """TR/EN normalization does not run for a string target, so the mix stays."""

    values = ["1.234,56", "1,234.56", "Kalem"]
    result = _result(
        {"note": values},
        {"source_file": ["a.csv"] * 3, "note": ["not"] * 3},
    )
    schema = _schema(("note", "string", False))
    mapping = _mapping(note=[("a.csv", "not", "auto")])

    finding = next(item for item in validate(result, mapping, schema).findings if item.check == "format")
    assert finding.severity == "warning"
    assert finding.count == 2


# --------------------------------------------------------------------------- #
# Required targets
# --------------------------------------------------------------------------- #


def test_a_required_column_with_a_null_is_an_error():
    result = _result(
        {"product_name": ["Kalem", None]},
        {"source_file": ["a.csv"] * 2, "product_name": ["urun"] * 2},
    )
    schema = _schema(("product_name", "string", True))
    mapping = _mapping(product_name=[("a.csv", "urun", "auto")])

    report = validate(result, mapping, schema)

    finding = next(item for item in report.findings if item.check == "required")
    assert finding.severity == "error"
    assert finding.count == 1
    assert report.blocking


def test_a_required_column_with_no_source_at_all_is_an_error():
    result = _result(
        {"product_name": [None, None]},
        {"source_file": ["a.csv"] * 2, "product_name": [None, None]},
    )
    schema = _schema(("product_name", "string", True))
    mapping = _mapping(product_name=[("a.csv", None, "unmatched")])

    finding = next(item for item in validate(result, mapping, schema).findings if item.check == "required")
    assert finding.severity == "error"
    assert "hiçbir kaynak dosyaya eşlenmedi" in finding.message


def test_a_clean_merge_produces_no_findings():
    result = _result(
        {"product_name": ["Kalem", "Defter"], "unit_price": [12.5, 8.9]},
        {
            "source_file": ["a.csv", "a.csv"],
            "product_name": ["urun", "urun"],
            "unit_price": ["fiyat", "fiyat"],
        },
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("product_name", "string", True), ("unit_price", "decimal", True))
    mapping = _mapping(
        product_name=[("a.csv", "urun", "auto")], unit_price=[("a.csv", "fiyat", "auto")]
    )

    report = validate(result, mapping, schema)

    assert report.findings == ()
    assert not report.blocking


# --------------------------------------------------------------------------- #
# The validator reports; it never repairs
# --------------------------------------------------------------------------- #


def test_validation_leaves_the_data_untouched():
    result = _result(
        {"unit_price": [12.5, None]},
        {"source_file": ["a.csv"] * 2, "unit_price": ["fiyat"] * 2},
        errors={"unit_price": 1},
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("unit_price", "decimal", True))
    mapping = _mapping(unit_price=[("a.csv", "fiyat", "auto")])
    before = result.dataframe.clone()

    report = validate(result, mapping, schema)

    assert report.blocking  # required column has a null
    assert result.dataframe.equals(before)
    assert result.dataframe["unit_price"].to_list() == [12.5, None]


def test_downgrade_to_review_touches_only_the_blamed_mapping():
    result = _result(
        {"product_name": ["Kalem", "Defter"], "unit_price": [None, None]},
        {
            "source_file": ["a.csv", "a.csv"],
            "product_name": ["urun", "urun"],
            "unit_price": ["fiyat", "fiyat"],
        },
        dtypes={"unit_price": pl.Float64},
    )
    schema = _schema(("product_name", "string", True), ("unit_price", "decimal", False))
    mapping = _mapping(
        product_name=[("a.csv", "urun", "auto")], unit_price=[("a.csv", "fiyat", "auto")]
    )

    downgraded = downgrade_to_review(mapping, validate(result, mapping, schema))

    statuses = {
        entry.target_column: [source.status for source in entry.sources]
        for entry in downgraded.entries
    }
    assert statuses == {"product_name": ["auto"], "unit_price": ["review"]}
    reason = downgraded.entries[1].sources[0].reason
    assert reason is not None and reason.startswith("validator: ")
    # The original plan object is not mutated.
    assert mapping.entries[1].sources[0].status == "auto"


def test_downgrade_leaves_a_clean_plan_alone():
    mapping = _mapping(product_name=[("a.csv", "urun", "auto")])
    result = _result(
        {"product_name": ["Kalem"]},
        {"source_file": ["a.csv"], "product_name": ["urun"]},
    )
    schema = _schema(("product_name", "string", True))

    assert downgrade_to_review(mapping, validate(result, mapping, schema)) is mapping


# --------------------------------------------------------------------------- #
# End to end through the CLI
# --------------------------------------------------------------------------- #


SCHEMA_YAML = """target_columns:
  - name: product_name
    type: string
    required: {product_required}
  - name: unit_price
    type: decimal
    required: false
  - name: stock_quantity
    type: integer
    required: false
output:
  format: csv
  add_provenance: true
"""

MAPPING_YAML = """- target_column: product_name
  sources:
    - file: sales.csv
      column: urun
      confidence: 0.99
      status: auto
- target_column: unit_price
  sources:
    - file: sales.csv
      column: fiyat
      confidence: 0.98
      status: auto
- target_column: stock_quantity
  sources:
    - file: sales.csv
      column: stok
      confidence: 0.97
      status: auto
"""


def _workspace(tmp_path: Path, rows: str, *, product_required: bool = False) -> Path:
    (tmp_path / "sales.csv").write_text("urun;fiyat;stok\n" + rows, encoding="utf-8")
    (tmp_path / "schema.yaml").write_text(
        SCHEMA_YAML.format(product_required=str(product_required).lower()), encoding="utf-8"
    )
    (tmp_path / "mapping.yaml").write_text(MAPPING_YAML, encoding="utf-8")
    return tmp_path


def _argv(workspace: Path, out: Path) -> list[str]:
    return ["apply", "--mapping", str(workspace / "mapping.yaml"), "--out", str(out), "--format", "csv"]


def test_apply_stops_and_writes_nothing_when_a_required_column_is_empty(tmp_path, capsys):
    workspace = _workspace(tmp_path, "Kalem;12,50;10\n;8,90;3\n", product_required=True)
    out = tmp_path / "merged.csv"

    exit_code = main(_argv(workspace, out))

    assert exit_code == REVIEW_GUARD_EXIT_CODE
    assert not out.exists()
    assert not (tmp_path / "merge_report.xlsx").exists()

    output = capsys.readouterr().out
    assert "validator" in output and "ciddi tutarsızlık" in output
    assert "[required] product_name" in output
    assert "product_name ← sales.csv:urun" in output
    assert "Veri değiştirilmedi" in output


def test_apply_stops_when_a_decimal_column_is_full_of_text(tmp_path, capsys):
    workspace = _workspace(tmp_path, "Kalem;abc;10\nDefter;def;3\nSilgi;ghi;1\n")
    out = tmp_path / "merged.csv"

    exit_code = main(_argv(workspace, out))

    assert exit_code == REVIEW_GUARD_EXIT_CODE
    assert not out.exists()
    output = capsys.readouterr().out
    assert "[type] unit_price ← sales.csv:fiyat" in output
    assert "unit_price ← sales.csv:fiyat" in output.split("review'a düşen")[1]


def test_warnings_reach_the_merge_report_without_stopping_the_merge(tmp_path, capsys):
    rows = "".join(f"Urun{index};{10 + index},50;\n" for index in range(9))
    workspace = _workspace(tmp_path, rows + "Urun9;abc;7\n")
    out = tmp_path / "merged.csv"

    exit_code = main(_argv(workspace, out))

    assert exit_code == 0
    assert pl.read_csv(out).height == 10  # no row is lost

    report = load_workbook(tmp_path / "merge_report.xlsx")
    summary = {row[0].value: row[1].value for row in report["Summary"].iter_rows(min_row=2)}
    assert summary["validation_errors"] == 0
    assert summary["validation_warnings"] >= 2

    findings = [
        {header.value: cell.value for header, cell in zip(report["Validation"][1], row)}
        for row in report["Validation"].iter_rows(min_row=2)
    ]
    checks = {finding["check"] for finding in findings}
    assert {"type", "null"} <= checks
    assert all(finding["severity"] == "warning" for finding in findings)
    assert any("null patlaması" in finding["message"] for finding in findings)

    output = capsys.readouterr().out
    assert "validator" in output and "uyarı" in output


def test_a_clean_run_reports_no_finding(tmp_path, capsys):
    workspace = _workspace(tmp_path, "Kalem;12,50;10\nDefter;8,90;3\n")
    out = tmp_path / "merged.csv"

    assert main(_argv(workspace, out)) == 0

    sheet = load_workbook(tmp_path / "merge_report.xlsx")["Validation"]
    assert [cell.value for cell in sheet[2]][1] == "ok"
    assert "tutarsızlık bulunamadı" in capsys.readouterr().out


def test_the_null_threshold_flag_is_honoured(tmp_path, capsys):
    workspace = _workspace(tmp_path, "Kalem;12,50;10\nDefter;8,90;\nSilgi;1,00;2\n")
    out = tmp_path / "merged.csv"

    assert main(_argv(workspace, out)) == 0
    assert "tutarsızlık bulunamadı" in capsys.readouterr().out

    assert main(_argv(workspace, out) + ["--null-threshold", "0.1"]) == 0
    assert "null patlaması" in capsys.readouterr().out


def test_the_validator_module_has_no_llm_dependency():
    """Spec §14: the second pair of eyes is deterministic, never a model."""

    import core.validator

    source = Path(core.validator.__file__).read_text(encoding="utf-8")
    assert "core.llm" not in source
    assert "llm" not in vars(core.validator)


@pytest.mark.parametrize("threshold", [-0.5, 1.5])
def test_an_out_of_range_threshold_is_rejected(threshold):
    with pytest.raises(Exception):
        ValidationSettings(null_warning_ratio=threshold)
