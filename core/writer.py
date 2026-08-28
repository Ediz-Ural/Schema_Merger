"""Deterministic writer for the merged table, provenance, and merge report.

The writer is the last deterministic step of Phase 3.  It takes the vertical
union produced by :mod:`core.transformer` and persists it as ``xlsx``, ``csv`` or
``sql`` while always appending provenance columns, and it emits a companion
``merge_report.xlsx``.

Design notes:

* **No LLM.**  Like the transformer, this module never touches :mod:`core.llm`.
* **Provenance is always added.**  ``schema.output.add_provenance`` is respected
  as a documented preference, but the invariant of the project (spec §14/§6) is
  that origin columns are *always* written so a merged row can be traced back to
  its source file and original column.  Passing ``add_provenance=False`` still
  writes them; the flag only records intent.
* **SQL is a script.**  The ``sql`` format produces a ``CREATE TABLE`` + ``INSERT``
  script (spec §2/§7).  No live database connection is ever opened.
* Two output files are kept clearly separate: ``merged.<fmt>`` holds the clean
  data and ``merge_report.xlsx`` holds the human-facing report.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Sequence

import polars as pl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.contracts import VALID_OUTPUT_FORMATS, ClusterContract, MappingContract, SchemaContract
from core.transformer import (
    MERGED_ROW_COUNT_COLUMN,
    DeduplicationResult,
    TransformationResult,
)
from core.types import EntityClusterPlan


#: Name of the provenance column carrying the originating source file per row.
SOURCE_FILE_COLUMN = "_source_file"

#: Default file name for the companion report placed next to the merged output.
DEFAULT_REPORT_NAME = "merge_report.xlsx"


class WriteError(ValueError):
    """Raised for an unsupported format or an inconsistent transformation."""


@dataclass(frozen=True)
class EntitySummary:
    """What entity resolution did to this merge, for the report (spec §5).

    ``clusters`` is the whole reviewed document, approved and pending alike: the
    report has to show the uncertain ones as well, because those products were
    deliberately *not* merged.
    """

    target_column: str
    clusters: tuple[EntityClusterPlan, ...] = ()
    merged_cluster_count: int = 0
    canonicalized_row_count: int = 0
    duplicate_row_count: int = 0

    @classmethod
    def from_deduplication(
        cls,
        dedup: DeduplicationResult,
        clusters: ClusterContract | Sequence[EntityClusterPlan],
    ) -> "EntitySummary":
        plans = clusters.clusters if isinstance(clusters, ClusterContract) else list(clusters)
        return cls(
            target_column=dedup.target_column,
            clusters=tuple(plans),
            merged_cluster_count=dedup.merged_cluster_count,
            canonicalized_row_count=dedup.canonicalized_row_count,
            duplicate_row_count=dedup.duplicate_row_count,
        )

    @property
    def pending_clusters(self) -> list[EntityClusterPlan]:
        """Clusters no one approved; their members stayed separate."""

        return [cluster for cluster in self.clusters if cluster.status == "review"]


@dataclass(frozen=True)
class WriteResult:
    """Where the two output files landed plus a short numeric summary."""

    merged_path: Path
    report_path: Path
    output_format: str
    row_count: int
    null_cell_count: int
    conversion_error_count: int


def source_column_name(target_column: str) -> str:
    """Provenance column name holding the original column for ``target_column``."""

    return f"{target_column}_source_column"


def write(
    result: TransformationResult,
    mapping: MappingContract,
    schema: SchemaContract,
    out_path: str | Path,
    *,
    output_format: str | None = None,
    report_path: str | Path | None = None,
    table_name: str | None = None,
    add_provenance: bool = True,
    entity: EntitySummary | None = None,
) -> WriteResult:
    """Write ``merged.<fmt>`` and ``merge_report.xlsx`` for one merge run.

    ``add_provenance`` only records the caller's intent: provenance columns are
    written regardless, matching the project invariant.  ``entity`` carries the
    approved-cluster outcome when entity resolution ran, so the report can show
    what merged and what stayed uncertain.
    """

    merged = Path(out_path)
    fmt = _resolve_format(output_format, schema, merged)

    merged_path = write_merged(result, schema, merged, output_format=fmt, table_name=table_name)
    report = Path(report_path) if report_path is not None else merged.parent / DEFAULT_REPORT_NAME
    report_out = write_merge_report(result, mapping, schema, report, output_format=fmt, entity=entity)

    return WriteResult(
        merged_path=merged_path,
        report_path=report_out,
        output_format=fmt,
        row_count=result.row_count,
        null_cell_count=_null_cell_count(result, schema),
        conversion_error_count=result.conversion_error_count,
    )


def write_merged(
    result: TransformationResult,
    schema: SchemaContract,
    out_path: str | Path,
    *,
    output_format: str | None = None,
    table_name: str | None = None,
) -> Path:
    """Write only the merged table (data + provenance) in the chosen format."""

    destination = Path(out_path)
    fmt = _resolve_format(output_format, schema, destination)
    combined = _combined_frame(result, schema)
    destination.parent.mkdir(parents=True, exist_ok=True)

    if fmt == "csv":
        combined.write_csv(destination)
    elif fmt == "xlsx":
        _write_frame_xlsx(combined, destination, sheet_title="merged")
    elif fmt == "sql":
        script = _build_sql_script(combined, result, schema, table_name or destination.stem)
        destination.write_text(script, encoding="utf-8", newline="\n")
    else:  # pragma: no cover - guarded by _resolve_format
        raise WriteError(f"Unsupported output format: {fmt!r}")
    return destination


def write_merge_report(
    result: TransformationResult,
    mapping: MappingContract,
    schema: SchemaContract,
    report_path: str | Path,
    *,
    output_format: str | None = None,
    entity: EntitySummary | None = None,
) -> Path:
    """Write ``merge_report.xlsx`` describing matches, row counts and nulls."""

    destination = Path(report_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _fill_summary_sheet(workbook, result, schema, output_format, entity)
    _fill_columns_sheet(workbook, result, mapping, schema)
    _fill_entity_sheet(workbook, entity)
    workbook.save(destination)
    return destination


# --------------------------------------------------------------------------- #
# Frame assembly
# --------------------------------------------------------------------------- #


def _combined_frame(result: TransformationResult, schema: SchemaContract) -> pl.DataFrame:
    """Merged target columns followed by their provenance columns.

    Column order is: every target column, then ``_source_file``, then one
    ``<target>_source_column`` per target column.
    """

    if result.dataframe.height != result.provenance.height:
        raise WriteError("Data and provenance frames must have the same number of rows")
    rename = {"source_file": SOURCE_FILE_COLUMN}
    for column in schema.target_columns:
        rename[column.name] = source_column_name(column.name)
    provenance = result.provenance.rename(rename)
    return result.dataframe.hstack(provenance)


def _resolve_format(output_format: str | None, schema: SchemaContract, out_path: Path) -> str:
    if output_format is None:
        output_format = schema.output.format
    fmt = output_format.lower()
    if fmt not in VALID_OUTPUT_FORMATS:
        raise WriteError(f"Unsupported output format {output_format!r}. Use one of: {', '.join(sorted(VALID_OUTPUT_FORMATS))}.")
    return fmt


def _null_cell_count(result: TransformationResult, schema: SchemaContract) -> int:
    return sum(result.dataframe[column.name].null_count() for column in schema.target_columns)


# --------------------------------------------------------------------------- #
# XLSX helpers
# --------------------------------------------------------------------------- #


def _write_frame_xlsx(frame: pl.DataFrame, path: Path, *, sheet_title: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append(list(frame.columns))
    for row in frame.iter_rows():
        # openpyxl writes None, str, numbers, bool, date and datetime natively.
        sheet.append(list(row))
    workbook.save(path)


def _fill_summary_sheet(
    workbook: Workbook,
    result: TransformationResult,
    schema: SchemaContract,
    output_format: str | None,
    entity: EntitySummary | None = None,
) -> None:
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["metric", "value"])
    rows = [
        ("merged_rows", result.row_count),
        ("target_columns", len(schema.target_columns)),
        ("total_null_cells", _null_cell_count(result, schema)),
        ("total_conversion_errors", result.conversion_error_count),
        ("output_format", output_format or schema.output.format),
        ("provenance_added", True),
    ]
    if entity is not None:
        rows.extend(
            [
                ("entity_column", entity.target_column),
                ("entity_clusters", len(entity.clusters)),
                ("entity_clusters_merged", entity.merged_cluster_count),
                ("entity_pending_clusters", len(entity.pending_clusters)),
                ("entity_canonicalized_rows", entity.canonicalized_row_count),
                ("entity_duplicate_rows_removed", entity.duplicate_row_count),
            ]
        )
    for metric, value in rows:
        sheet.append([metric, value])
    _autosize(sheet)


def _fill_columns_sheet(
    workbook: Workbook,
    result: TransformationResult,
    mapping: MappingContract,
    schema: SchemaContract,
) -> None:
    sheet = workbook.create_sheet("Columns")
    headers = [
        "target_column",
        "type",
        "required",
        "source_file",
        "source_column",
        "confidence",
        "status",
        "reason",
        "null_count",
        "null_ratio",
        "conversion_errors",
    ]
    sheet.append(headers)

    entries = {entry.target_column: entry for entry in mapping.entries}
    row_count = result.row_count
    for column in schema.target_columns:
        null_count = result.dataframe[column.name].null_count()
        null_ratio = round(null_count / row_count, 6) if row_count else 0.0
        errors = result.conversion_error_counts.get(column.name, 0)
        sources = entries[column.name].sources if column.name in entries else []
        if not sources:
            sheet.append(
                [column.name, column.type, column.required, None, None, None, None, None, null_count, null_ratio, errors]
            )
            continue
        for source in sources:
            sheet.append(
                [
                    column.name,
                    column.type,
                    column.required,
                    source.file,
                    source.column,
                    round(source.confidence, 6),
                    source.status,
                    source.reason,
                    null_count,
                    null_ratio,
                    errors,
                ]
            )
    _autosize(sheet)


def _fill_entity_sheet(workbook: Workbook, entity: EntitySummary | None = None) -> None:
    """Report entity clusters: what merged, and what stayed uncertain (spec §5).

    Pending clusters are written too, and marked plainly: those products were
    *not* deduplicated because nobody approved them.
    """

    sheet = workbook.create_sheet("Entity")
    sheet.append(
        [
            "target_column",
            "cluster_id",
            "representative",
            "member_count",
            "row_count",
            "status",
            "members",
            "candidates",
            "note",
        ]
    )
    if entity is None:
        sheet.append(
            [
                None,
                None,
                None,
                None,
                None,
                "not_run",
                None,
                None,
                "Entity resolution bu çalıştırmada koşmadı; --clusters ile bağlanır.",
            ]
        )
        _autosize(sheet)
        return
    if not entity.clusters:
        sheet.append(
            [
                entity.target_column,
                None,
                None,
                None,
                None,
                "empty",
                None,
                None,
                "Küme dosyası boş; hiçbir değer birleştirilmedi.",
            ]
        )
        _autosize(sheet)
        return

    for cluster in entity.clusters:
        sheet.append(
            [
                cluster.target_column or entity.target_column,
                cluster.cluster_id,
                cluster.canonical,
                len(cluster.members),
                sum(member.row_count for member in cluster.members),
                cluster.status,
                "; ".join(member.value for member in cluster.members),
                "; ".join(candidate.value for candidate in cluster.candidates),
                _entity_note(cluster),
            ]
        )
    _autosize(sheet)


def _entity_note(cluster: EntityClusterPlan) -> str:
    if cluster.status == "review":
        return "BELİRSİZ: onaylanmadı, üyeler birleştirilmedi. " + (cluster.reason or "")
    if cluster.status == "rejected":
        return "Kullanıcı reddetti: farklı ürünler. " + (cluster.reason or "")
    return cluster.reason or "Onaylı küme; canonical değere getirildi."


def _autosize(sheet) -> None:
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 60)


# --------------------------------------------------------------------------- #
# SQL helpers
# --------------------------------------------------------------------------- #


_SQL_TYPES = {
    "string": "TEXT",
    "integer": "INTEGER",
    "decimal": "REAL",
    "date": "DATE",
    "boolean": "BOOLEAN",
}
_IDENTIFIER_CLEAN = re.compile(r"\W+")


def _build_sql_script(
    frame: pl.DataFrame,
    result: TransformationResult,
    schema: SchemaContract,
    table_name: str,
) -> str:
    """Build a ``CREATE TABLE`` + ``INSERT`` script (no live DB, spec §2/§7)."""

    table = _sql_identifier(table_name or "merged")
    sql_types = _column_sql_types(frame, schema)

    lines: list[str] = []
    lines.append(f"-- Schema Merger output: {result.row_count} row(s), provenance columns included.")
    lines.append(f"DROP TABLE IF EXISTS {table};")
    column_defs = ",\n".join(f"    {_sql_identifier(name)} {sql_types[name]}" for name in frame.columns)
    lines.append(f"CREATE TABLE {table} (\n{column_defs}\n);")

    column_list = ", ".join(_sql_identifier(name) for name in frame.columns)
    for row in frame.iter_rows():
        values = ", ".join(_sql_literal(value) for value in row)
        lines.append(f"INSERT INTO {table} ({column_list}) VALUES ({values});")
    return "\n".join(lines) + "\n"


def _column_sql_types(frame: pl.DataFrame, schema: SchemaContract) -> dict[str, str]:
    types = {column.name: _SQL_TYPES[column.type] for column in schema.target_columns}
    # Provenance columns are text, except the merged-row counter.
    types.setdefault(MERGED_ROW_COUNT_COLUMN, "INTEGER")
    for name in frame.columns:
        types.setdefault(name, "TEXT")
    return types


def _sql_identifier(name: str) -> str:
    cleaned = _IDENTIFIER_CLEAN.sub("_", name.strip()) or "col"
    if cleaned[0].isdigit():
        cleaned = f"_{cleaned}"
    return f'"{cleaned}"'


def _sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        return repr(value)
    if isinstance(value, datetime):
        return "'" + value.isoformat(sep=" ") + "'"
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    return "'" + str(value).replace("'", "''") + "'"
